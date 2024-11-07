import os
import numpy as np
import torch
import torch.nn as nn
from PIL import Image
from tqdm import tqdm
from src.utils.model_io import load_weights
from src.dataloader.zjuL5 import ZJUL5

from src.utils.utils import RunningAverageDict, colorize,make_model
from src.config import args
from src.dataloader.nyu import NYUV2
from PIL import Image, ImageFont, ImageDraw, ImageEnhance
from torch.distributions import categorical
def compute_errors(gt, pred):
    thresh = np.maximum((gt / pred), (pred / gt))
    a1 = (thresh < 1.25).mean()
    a2 = (thresh < 1.25 ** 2).mean()
    a3 = (thresh < 1.25 ** 3).mean()

    abs_rel = np.mean(np.abs(gt - pred) / gt)
    sq_rel = np.mean(((gt - pred) ** 2) / gt)

    rmse = (gt - pred) ** 2
    rmse = np.sqrt(rmse.mean())

    rmse_log = (np.log(gt) - np.log(pred)) ** 2
    rmse_log = np.sqrt(rmse_log.mean())

    err = np.log(pred) - np.log(gt)
    silog = np.sqrt(np.mean(err ** 2) - np.mean(err) ** 2) * 100

    log_10 = (np.abs(np.log10(gt) - np.log10(pred))).mean()
    return dict(a1=a1, a2=a2, a3=a3, abs_rel=abs_rel, rmse=rmse,
                log_10=log_10, rmse_log=rmse_log, silog=silog, sq_rel=sq_rel)

def predict_tta(model, input_data, args):

    pred_refine,pred, prob, prob_residual = model(input_data)
    pred = np.clip(pred.cpu().numpy(), args.min_depth, args.max_depth)
    pred = nn.functional.interpolate(torch.from_numpy(pred), input_data['rgb'].shape[-2:], mode='bilinear', align_corners=True)
    prob = nn.functional.interpolate(prob, input_data['rgb'].shape[-2:], mode='bilinear', align_corners=True)

    return pred,  prob

def eval(model, test_loader, args, device):
    if args.save_dir is not None and not os.path.exists(f'{args.save_dir}'):
        os.system(f'mkdir -p {args.save_dir}')

    metrics = RunningAverageDict()
    with torch.no_grad():
        model.eval()
        for index, batch in enumerate(tqdm(test_loader)):
            gt = batch['depth'].to(device)
            img = batch['image'].to(device)
            input_data = {'rgb': img}
            additional_data = {}
            additional_data['hist_data'] = batch['additional']['hist_data'].to(device)
            additional_data['rect_data'] = batch['additional']['rect_data'].to(device)
            additional_data['mask'] = batch['additional']['mask'].to(device)
            additional_data['patch_info'] = batch['additional']['patch_info']
            input_data.update({
                'additional': additional_data
            })

            final,prob = predict_tta(model, input_data, args)
            final = final.squeeze().cpu().numpy()
            if 'Deltar_rgb' == model._get_name():
                final *= np.median(np.median(batch['additional']['hist_data'][0].mean(-1)[batch['additional']['mask'][0]>0]))/np.median(final)

            impath = f"{batch['image_path'][0].replace('/', '__').replace('.jpg', '')}"
            im_subfolder = batch['image_folder'][0]
            vis_folder = f'{im_subfolder}'
            if not os.path.exists(f'{args.save_dir}/{vis_folder}'):
                os.system(f'mkdir -p {args.save_dir}/{vis_folder}')


            gt = gt.squeeze().cpu().numpy()

            valid_mask = np.logical_and(gt > args.min_depth, gt < args.max_depth)
            # valid_mask = np.logical_and(valid_mask, gt<4.)
            # valid_mask = np.logical_and(valid_mask, gt>3.)
            if valid_mask.sum()>0:
                metrics.update(compute_errors(gt[valid_mask], final[valid_mask]))

            # break

    metrics = {k: round(v, 3) for k, v in metrics.get_value().items()}
    print(f"Metrics: {metrics}")
    print(','.join([str(i) for i in metrics.values()]))
    return metrics

if __name__ == '__main__':

    device = torch.device('cuda:0')

    if 'nyu' in args.test_dataset:
        pass
    elif 'zjuL5' in args.test_dataset:
        args.data_path_eval = 'data/ZJUL5'
        args.filenames_file_eval = 'data/ZJUL5/data.json'
        args.input_height = 480
        args.input_width = 640
        args.max_depth = 10
        args.min_depth = 1e-3
        args.n_bins = 256
        args.min_depth_eval = 1e-3
        args.max_depth_eval = 10
        args.zone_sample_num = 16
    else:
        raise NotImplementedError

    import openpyxl

    # 获取 工作簿对象

    workbook = openpyxl.Workbook()

    worksheet = workbook.active
    # worksheet.title = "results"

    for i,metric in enumerate(['epoch','a1',  'a2', 'a3', 'abs_rel', 'rmse', 'log_10',  'rmse_log', 'silog', 'sq_rel']):
        worksheet.cell(1,i+1,metric)



    start = 0
    end = args.epochs

    for ep in range(start,end):
        print(ep)
        if 'ZJU' in args.filenames_file_eval:
            test_loader = ZJUL5(args, 'online_eval').data
            dataset = 'ZJUL5'
        elif 'nyu' in args.filenames_file_eval:
            test_loader = NYUV2(args, 'online_eval').data
            dataset = 'NYUv2'

        model=make_model(args)
        model = model.to(device)

        if args.selected_epoch !='-1':
            here_name = args.selected_epoch
            weight_path = os.path.join('weights', args.name, f"{here_name}.pt")
        else:
            possible_names = os.listdir(os.path.join('weights',args.name))
            here_name = [i for i in possible_names if i.startswith(f'{ep}_')]
            assert len(here_name)<=1
            if len(here_name) ==0:
                continue
            here_name = here_name[0]
            weight_path = os.path.join('weights',args.name, f"{here_name}")

        model = load_weights(model,weight_path)
        model = model.eval()

        results = eval(model, test_loader, args, device)
        worksheet.cell(ep + 2, 1, ep)
        for i, metric in enumerate(['a1', 'a2', 'a3', 'abs_rel', 'rmse', 'log_10', 'rmse_log','silog', 'sq_rel']):
            worksheet.cell(ep+2, i + 2, results[metric])
        if args.selected_epoch != '-1':
            break
        # break
    if 'nyu' in args.test_dataset:
        workbook.save(filename=os.path.join(args.save_dir,'results_nyu.xlsx'))
    else:
        workbook.save(filename=os.path.join(args.save_dir,'results.xlsx'))